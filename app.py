import streamlit as st
import pandas as pd
import os
from pathlib import Path
from datetime import datetime, date
import openpyxl
import database as db

# ── Cookie helpers (persistenza F5 senza esporre la sessione nella URL) ──
def _set_cookie(name, value, max_age=604800):
    st.markdown(f"""<script>document.cookie="{name}={value};path=/;max-age={max_age};SameSite=Lax;Secure";</script>""", unsafe_allow_html=True)

def _del_cookie(name):
    st.markdown(f"""<script>document.cookie="{name}=;path=/;max-age=0;SameSite=Lax;Secure";</script>""", unsafe_allow_html=True)

def _cookie_to_url(cname="sygs_sid", pname="_sid"):
    """Se il cookie esiste, lo trasferisce in query param e ricarica."""
    st.markdown(f"""<script>
(function(){{
    const m=document.cookie.match(new RegExp('(^| ){cname}=([^;]+)'));
    if(m && m[2] && !window.location.search.includes('{pname}=')){{
        const u=new URL(window.location);
        u.searchParams.set('{pname}',m[2]);
        window.history.replaceState({{}},'',u);
        window.location.reload();
    }}
}})();
</script>""", unsafe_allow_html=True)

def _remove_url_param(pname="_sid"):
    """Rimuove un query param dalla barra indirizzi via JS."""
    st.markdown(f"""<script>
(function(){{
    const u=new URL(window.location);
    u.searchParams.delete('{pname}');
    window.history.replaceState({{}},'',u);
}})();
</script>""", unsafe_allow_html=True)

# ── Carica .env se esiste (per esecuzione locale) ──
_env_path = Path(__file__).parent / ".env"
if _env_path.exists():
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip().strip("\"'"))


def leggi_excel_con_formule(source, sheet=0):
    """Legge un Excel preservando le formule come testo (=...) invece dei valori calcolati.
    source può essere un file-like object (upload) o un percorso di file."""
    import io
    if isinstance(source, (str, Path)):
        with open(source, 'rb') as f:
            buf = io.BytesIO(f.read())
    else:
        source.seek(0)
        buf = io.BytesIO(source.read())
    buf.seek(0)
    df = pd.read_excel(buf, sheet_name=sheet, engine='openpyxl')
    if df.empty:
        return df
    buf.seek(0)
    wb = openpyxl.load_workbook(buf, data_only=False)
    ws = wb.active
    if ws.max_row < 2:
        wb.close()
        return df
    headers = [cell.value for cell in ws[1]]
    col_map = {}
    for ci, h in enumerate(headers):
        if h is not None and h in df.columns:
            col_map[ci] = h
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        excel_row = row[0].row
        df_row = excel_row - 2
        if df_row >= len(df):
            break
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                ci = cell.column - 1
                if ci in col_map:
                    df.at[df_row, col_map[ci]] = cell.value
    wb.close()
    return df

st.set_page_config(page_title="SYGS MASTER VIEWER", page_icon="📊", layout="wide",
                   initial_sidebar_state="expanded")

# ── CSS ─────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .main > div { padding: 0.5rem 1.5rem; }
    div[data-testid="stSidebarContent"] { padding: 1rem 0.8rem; }

    .login-box {
        max-width: 400px; margin: 4rem auto; padding: 2.5rem;
        background: white; border-radius: 16px; box-shadow: 0 4px 24px rgba(0,0,0,0.08);
        text-align: center;
    }
    .login-box h1 { font-size: 1.6rem; font-weight: 700; color: #1E3A5F; margin-bottom: 0.2rem; }
    .login-box .subtitle { color: #6B7280; font-size: 0.85rem; margin-bottom: 1.5rem; }

    .card {
        background: white; border-radius: 12px; padding: 1.2rem 1.5rem;
        box-shadow: 0 1px 6px rgba(0,0,0,0.04); border: 1px solid #f0f0f0;
        margin-bottom: 0.6rem; transition: 0.15s;
    }
    .card:hover { border-color: #1E3A5F20; box-shadow: 0 2px 12px rgba(30,58,95,0.06); }

    .badge-online {
        display: inline-block; width: 8px; height: 8px; border-radius: 50%;
        background: #10B981; margin-right: 6px;
    }
    .badge-offline {
        display: inline-block; width: 8px; height: 8px; border-radius: 50%;
        background: #D1D5DB; margin-right: 6px;
    }

    .page-title { font-size: 1.5rem; font-weight: 700; color: #1E3A5F; margin-bottom: 0.3rem; }
    .page-sub { color: #6B7280; font-size: 0.9rem; margin-bottom: 1.5rem; }

    div.stButton > button { border-radius: 8px; font-weight: 500; }
    div.stButton > button[data-kind="primary"] { background: #1E3A5F; color: white; }
    div[data-testid="stDataFrame"] { border: 1px solid #e5e7eb; border-radius: 8px; overflow: hidden; }
    div[data-testid="stVerticalBlockBorderWrapper"] > div { gap: 4px; }
    .kanban-card { background: white; border: 1px solid #e5e7eb; border-radius: 8px; padding: 6px 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.04); }


    .user-avatar {
        width: 36px; height: 36px; border-radius: 50%; display: inline-flex;
        align-items: center; justify-content: center; color: white;
        font-weight: 600; font-size: 0.9rem; flex-shrink: 0;
    }
    .user-row { display: flex; align-items: center; gap: 8px; margin-bottom: 4px; }
</style>
""", unsafe_allow_html=True)

# ── Session State ───────────────────────────────────────────────
INIT = {
    "pagina": "login", "utente": None, "msg": "", "msg_tipo": "success",
    "file_id": None, "filtro_col": "", "filtro_val": "", "ricerca": "",
    "modifica_id": None, "vedi_id": None, "visuale": "tabella",
    "modifiche_pendenti": [], "formula_mode": False,
}
for k, v in INIT.items():
    if k not in st.session_state:
        st.session_state[k] = v


def msg(text, tipo="success"):
    st.session_state.msg = text
    st.session_state.msg_tipo = tipo


def cambia_pagina(pag):
    st.session_state.pagina = pag
    st.session_state.msg = ""
    st.rerun()


def iniziali(nome):
    parts = nome.strip().split()
    return (parts[0][0] + (parts[-1][0] if len(parts) > 1 else "")).upper()


try:
    ADMIN_EMAIL = st.secrets["ADMIN_EMAIL"]
    ADMIN_NOME = st.secrets["ADMIN_NOME"]
    ADMIN_PASSWORD = st.secrets["ADMIN_PASSWORD"]
except (KeyError, AttributeError, Exception):
    ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL")
    ADMIN_NOME = os.environ.get("ADMIN_NOME")
    ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")

if not ADMIN_EMAIL or not ADMIN_PASSWORD:
    st.markdown("""<div style="max-width:600px;margin:4rem auto;padding:2rem;background:#fff3f3;border-radius:12px;border:1px solid #e74c3c;">
<h2 style="color:#e74c3c;margin:0 0 0.5rem 0;">⚙️ Configurazione necessaria</h2>
<p style="color:#555;">Crea un file <code>.env</code> nella cartella del progetto:</p>
<pre style="background:#f5f5f5;padding:0.8rem;border-radius:6px;font-size:0.85rem;">
ADMIN_EMAIL = la tua email
ADMIN_NOME = Il tuo nome
ADMIN_PASSWORD = la tua password
</pre>
<p style="color:#555;">Oppure su <b>Streamlit Cloud</b> → ⚙️ Settings → Secrets:</p>
<pre style="background:#f5f5f5;padding:0.8rem;border-radius:6px;font-size:0.85rem;">
ADMIN_EMAIL = "la tua email"
ADMIN_NOME = "Il tuo nome"
ADMIN_PASSWORD = "la tua password"
</pre></div>""", unsafe_allow_html=True)
    st.stop()
db.init_db(admin_email=ADMIN_EMAIL, admin_nome=ADMIN_NOME, admin_password=ADMIN_PASSWORD)

# ── Ripristina sessione (cookie → URL sicura, sopravvive a F5) ──
if not st.session_state.utente:
    if "_sid" in st.query_params:
        sid = st.query_params["_sid"]
        del st.query_params["_sid"]
        if sid:
            utente = db.leggi_sessione(sid)
            if utente:
                st.session_state.utente = utente
                st.session_state.sid = sid
                st.session_state.pagina = "dashboard"
        _remove_url_param("_sid")
    else:
        # Se il cookie esiste, JS lo copia in ?_sid= e ricarica
        # Se non esiste, si prosegue verso il login
        _cookie_to_url()

# ── SIDEBAR ─────────────────────────────────────────────────────
with st.sidebar:
    if st.session_state.utente:
        u = st.session_state.utente
        col_av, col_info = st.columns([1, 3])
        with col_av:
            st.markdown(
                f'<div class="user-avatar" style="background:{u["avatar_color"]}">{iniziali(u["nome"])}</div>',
                unsafe_allow_html=True,
            )
        with col_info:
            st.markdown(f"**{u['nome']}**  ")
            st.caption(f"{u['email']}  ·  {'Admin' if u['ruolo']=='admin' else 'Utente'}")

        st.divider()

        # Navigation
        nav_items = [
            ("🏠", "Dashboard", "dashboard"),
            ("📂", "Carica Excel", "carica"),
        ]
        if u["ruolo"] == "admin":
            nav_items.append(("👥", "Gestione Utenti", "utenti"))

        for icon, label, page in nav_items:
            if st.button(f"{icon}  {label}", use_container_width=True, type="secondary" if st.session_state.pagina != page else "primary"):
                cambia_pagina(page)

        st.divider()

        # File list
        files = db.file_utente(u["id"])
        if files:
            st.markdown("**📁 File aperti**")
            for f in files[:5]:
                cols = st.columns([1, 4])
                with cols[0]:
                    is_active = st.session_state.file_id == f["id"]
                    st.markdown("📊" if not is_active else "📌")
                with cols[1]:
                    if st.button(f["nome_file"], key=f"side_f_{f['id']}", help=os.path.dirname(f["percorso"]),
                                 use_container_width=True, type="secondary" if not is_active else "primary"):
                        st.session_state.file_id = f["id"]
                        st.session_state.pagina = "vedi_file"
                        st.rerun()

        st.divider()
        with st.expander("🔑 Cambia password"):
            with st.form("cambia_mia_pw", border=False):
                old = st.text_input("Password attuale", type="password", key="old_pw")
                new1 = st.text_input("Nuova password", type="password", key="new_pw1")
                if st.form_submit_button("Aggiorna", use_container_width=True):
                    if old and new1:
                        ut = db.login(u["email"], old)
                        if ut:
                            db.cambia_password(u["id"], new1)
                            msg("Password cambiata")
                            st.rerun()
                        else:
                            msg("Password attuale errata", "error")
                    else:
                        msg("Compila tutti i campi", "warning")
        st.divider()
        if st.button("🚪  Esci", use_container_width=True):
            sid = st.session_state.get("sid")
            if sid:
                db.elimina_sessione(sid)
            _del_cookie("sygs_sid")
            _remove_url_param("_sid")
            st.session_state.utente = None
            st.session_state.pagina = "login"
            st.session_state.file_id = None
            st.rerun()

# ── MAIN ─────────────────────────────────────────────────────────
if st.session_state.msg:
    fn = getattr(st, st.session_state.msg_tipo)
    fn(st.session_state.msg)
    st.session_state.msg = ""

# ── FORMULA ENGINE ────────────────────────────────────────────
import re as _re
import math as _math

class _FormulaError(Exception):
    pass

def _col_letter_to_idx(letters: str) -> int:
    r = 0
    for ch in letters.upper():
        r = r * 26 + (ord(ch) - 64)
    return r - 1

def _tokenize(expr: str):
    tokens = []
    i = 0
    n = len(expr)
    while i < n:
        ch = expr[i]
        if ch.isspace():
            i += 1
        elif ch.isdigit() or (ch == '.' and i+1 < n and expr[i+1].isdigit()):
            start = i
            while i < n and (expr[i].isdigit() or expr[i] == '.'):
                i += 1
            tokens.append(('num', float(expr[start:i])))
        elif ch.isalpha() or ch == '_':
            start = i
            while i < n and (expr[i].isalnum() or expr[i] == '_'):
                i += 1
            word = expr[start:i]
            if i < n and expr[i] == '(':
                tokens.append(('func', word.upper()))
            else:
                tokens.append(('cell', word))
        elif ch in '+-*/^%()=,<>':
            if ch in '<>=' and i+1 < n and expr[i+1] == '=':
                tokens.append(('op', expr[i:i+2]))
                i += 2
            elif ch == '=':
                tokens.append(('op', '='))
                i += 1
            elif ch == '<':
                tokens.append(('op', '<'))
                i += 1
            elif ch == '>':
                tokens.append(('op', '>'))
                i += 1
            else:
                tokens.append(('op' if ch in '+-*/^%' else 'paren' if ch in '()' else 'comma', ch))
                i += 1
        elif ch in '"\'':
            q = ch; i += 1; start = i
            while i < n and expr[i] != q: i += 1
            tokens.append(('str', expr[start:i]))
            i += 1
        else:
            i += 1
    tokens.append(('eof', None))
    return tokens

class _FP:
    def __init__(self, toks, dcols, df, crow):
        self.toks = toks; self.pos = 0
        self.dcols = dcols; self.df = df; self.crow = crow
    def peek(self):
        return self.toks[self.pos] if self.pos < len(self.toks) else ('eof', None)
    def eat(self, t=None, v=None):
        p = self.peek()
        if t and p[0] != t: raise _FormulaError(f"Aspettato {t}, ottenuto {p[0]}")
        if v and p[1] != v: raise _FormulaError(f"Aspettato '{v}', ottenuto '{p[1]}'")
        self.pos += 1; return p
    def parse(self):
        return self._expr()
    def _expr(self):
        r = self._comp()
        while self.peek()[0] == 'op' and self.peek()[1] in ('+', '-'):
            op = self.eat()[1]; right = self._comp()
            r = (r + right) if op == '+' else (r - right)
        return r
    def _comp(self):
        r = self._term()
        while self.peek()[0] == 'op' and self.peek()[1] in ('=', '>', '<', '>=', '<=', '<>'):
            op = self.eat()[1]; right = self._term()
            if op == '=': r = r == right
            elif op == '>': r = r > right
            elif op == '<': r = r < right
            elif op == '>=': r = r >= right
            elif op == '<=': r = r <= right
            elif op == '<>': r = r != right
        return r
    def _term(self):
        r = self._pow()
        while self.peek()[0] == 'op' and self.peek()[1] in ('*', '/', '%'):
            op = self.eat()[1]; right = self._pow()
            if op == '*': r = r * right
            elif op == '/': r = r / right if right != 0 else '#DIV/0!'
            elif op == '%': r = r % right if right != 0 else '#DIV/0!'
        return r
    def _pow(self):
        r = self._unary()
        if self.peek()[0] == 'op' and self.peek()[1] == '^':
            self.eat(); r = r ** self._unary()
        return r
    def _unary(self):
        if self.peek()[0] == 'op' and self.peek()[1] == '-':
            self.eat(); return -self._prim()
        if self.peek()[0] == 'op' and self.peek()[1] == '+':
            self.eat(); return self._prim()
        return self._prim()
    def _cell_val(self, ref):
        if ':' in ref:
            return self._range_val(ref)
        m = _re.match(r'^([A-Za-z]+)(\d+)$', ref)
        if m:
            ci = _col_letter_to_idx(m.group(1))
            ri = int(m.group(2)) - 1
            if 0 <= ci < len(self.dcols) and 0 <= ri < len(self.df):
                v = self.df.iloc[ri][self.dcols[ci]]
                if pd.notna(v):
                    try: return float(v)
                    except: return v
            return 0
        return 0
    def _range_val(self, ref):
        parts = ref.split(':')
        if len(parts) != 2: return []
        m1 = _re.match(r'^([A-Za-z]+)(\d+)$', parts[0])
        m2 = _re.match(r'^([A-Za-z]+)(\d+)$', parts[1])
        if not m1 or not m2: return []
        c1, r1 = _col_letter_to_idx(m1.group(1)), int(m1.group(2)) - 1
        c2, r2 = _col_letter_to_idx(m2.group(1)), int(m2.group(2)) - 1
        vals = []
        for r in range(min(r1,r2), max(r1,r2)+1):
            for c in range(min(c1,c2), max(c1,c2)+1):
                if 0 <= c < len(self.dcols) and 0 <= r < len(self.df):
                    v = self.df.iloc[r][self.dcols[c]]
                    if pd.notna(v):
                        try: vals.append(float(v))
                        except: pass
        return vals
    def _func(self, name, args):
        if name == 'SUM':
            vals = [v for a in args for v in (a if isinstance(a, list) else [a]) if isinstance(v, (int, float))]
            return sum(vals)
        elif name in ('AVERAGE','AVG'):
            vals = [v for a in args for v in (a if isinstance(a, list) else [a]) if isinstance(v, (int, float))]
            return sum(vals)/len(vals) if vals else 0
        elif name == 'COUNT':
            return sum(1 for a in args for v in (a if isinstance(a, list) else [a]) if v is not None and v != '')
        elif name == 'MIN':
            vals = [v for a in args for v in (a if isinstance(a, list) else [a]) if isinstance(v, (int, float))]
            return min(vals) if vals else 0
        elif name == 'MAX':
            vals = [v for a in args for v in (a if isinstance(a, list) else [a]) if isinstance(v, (int, float))]
            return max(vals) if vals else 0
        elif name == 'IF':
            if len(args) >= 2: return args[1] if args[0] else (args[2] if len(args) > 2 else '')
            return 0
        elif name == 'ABS':
            return abs(args[0]) if args else 0
        elif name == 'ROUND':
            if len(args) >= 2: return round(args[0], int(args[1]))
            return round(args[0]) if args else 0
        elif name == 'PI':
            return _math.pi
        elif name == 'NOW':
            from datetime import datetime
            return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elif name == 'TODAY':
            from datetime import date
            return str(date.today())
        elif name == 'UPPER':
            return str(args[0]).upper() if args else ''
        elif name == 'LOWER':
            return str(args[0]).lower() if args else ''
        elif name == 'TRIM':
            return str(args[0]).strip() if args else ''
        elif name == 'LEN':
            return len(str(args[0])) if args else 0
        return 0
    def _prim(self):
        t = self.peek()
        if t[0] == 'num': self.eat(); return t[1]
        if t[0] == 'str': self.eat(); return t[1]
        if t[0] == 'cell': self.eat(); return self._cell_val(t[1])
        if t[0] == 'func':
            name = self.eat()[1]
            self.eat('paren', '(')
            args = []
            while not (self.peek()[0] == 'paren' and self.peek()[1] == ')'):
                if args: self.eat('comma')
                args.append(self._expr())
            self.eat('paren', ')')
            return self._func(name, args)
        if t[0] == 'paren' and t[1] == '(':
            self.eat(); r = self._expr(); self.eat('paren', ')'); return r
        return 0

def evaluar_formula(formula, df, row_idx, data_cols):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    expr = formula[1:].strip()
    if not expr: return ''
    try:
        toks = _tokenize(expr)
        r = _FP(toks, data_cols, df, row_idx).parse()
        if isinstance(r, float):
            return int(r) if r == int(r) else round(r, 2)
        return r
    except Exception:
        return '#ERROR!'

def aplicar_formulas(df, data_cols):
    if not st.session_state.get("formula_mode"):
        return df
    df2 = df.copy()
    for col in data_cols:
        for idx in df2.index:
            val = df2.at[idx, col]
            if isinstance(val, str) and val.startswith('='):
                df2.at[idx, col] = evaluar_formula(val, df, idx, data_cols)
    return df2

# ── LOGIN ────────────────────────────────────────────────────────
if st.session_state.pagina == "login":
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown('<div class="login-box">', unsafe_allow_html=True)
        st.markdown('<div style="font-size:2.5rem;margin-bottom:0.3rem;">📊</div>', unsafe_allow_html=True)
        st.markdown('<h1>SYGS MASTER VIEWER</h1>', unsafe_allow_html=True)
        st.markdown('<p class="subtitle">Gestione dati Excel multi-utente</p>', unsafe_allow_html=True)
        email = st.text_input("Email", placeholder="nome@esempio.com")
        password = st.text_input("Password", type="password", placeholder="••••••••")
        if st.button("Accedi", use_container_width=True, type="primary"):
            if email and password:
                utente = db.login(email, password)
                if utente:
                    sid = db.crea_sessione(utente)
                    _set_cookie("sygs_sid", sid)
                    st.session_state.utente = utente
                    st.session_state.sid = sid
                    st.session_state.pagina = "dashboard"
                    msg(f"Benvenuto, {utente['nome']}!")
                    st.rerun()
                else:
                    msg("Email o password non validi", "error")
            else:
                msg("Inserisci email e password", "warning")
        st.markdown('<div style="margin-top:1rem;font-size:0.8rem;color:#9CA3AF;">'
                    'Primo accesso: usa le credenziali fornite dall\'amministratore</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

# ── Guard ────────────────────────────────────────────────────────
if not st.session_state.utente:
    st.session_state.pagina = "login"
    st.rerun()

db.aggiorna_accesso(st.session_state.utente["id"])

# ── DASHBOARD ───────────────────────────────────────────────────
if st.session_state.pagina == "dashboard":
    u = st.session_state.utente
    st.markdown(f'<div class="page-title">🏠 Dashboard</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="page-sub">Benvenuto, {u["nome"]}</div>', unsafe_allow_html=True)

    # Stats
    files = db.file_utente(u["id"])
    online = db.utenti_online()
    col_s1, col_s2, col_s3 = st.columns(3)
    with col_s1:
        st.markdown(f'<div class="card"><h3 style="margin:0;color:#1E3A5F;">{len(files)}</h3>'
                    f'<div style="color:#6B7280;font-size:0.85rem;">File Excel caricati</div></div>',
                    unsafe_allow_html=True)
    with col_s2:
        tot = sum(f["righe"] for f in files)
        st.markdown(f'<div class="card"><h3 style="margin:0;color:#1E3A5F;">{tot}</h3>'
                    f'<div style="color:#6B7280;font-size:0.85rem;">Record totali</div></div>',
                    unsafe_allow_html=True)
    with col_s3:
        st.markdown(f'<div class="card"><h3 style="margin:0;color:#1E3A5F;">{len(online)}</h3>'
                    f'<div style="color:#6B7280;font-size:0.85rem;">Utenti online</div></div>',
                    unsafe_allow_html=True)

    st.divider()

    # Online users
    st.markdown("#### 👥 Utenti online")
    if online:
        for ou in online:
            col_img, col_nome = st.columns([0.5, 3])
            with col_img:
                st.markdown(
                    f'<div class="user-avatar" style="background:{ou["avatar_color"]}">{iniziali(ou["nome"])}</div>',
                    unsafe_allow_html=True,
                )
            with col_nome:
                st.markdown(f"**{ou['nome']}**  ·  {ou['email']}")
    else:
        st.caption("Nessun altro utente online")

    st.divider()

    # Recent files
    st.markdown("#### 📁 I tuoi file")
    if files:
        for f in files:
            cols = st.columns([2.5, 1, 1, 1.2, 0.5])
            with cols[0]:
                st.markdown(f"**{f['nome_file']}**")
                st.caption(os.path.dirname(f["percorso"]))
            with cols[1]:
                st.markdown(f"📄 {f['foglio']}")
            with cols[2]:
                st.markdown(f"{f['righe']} righe")
            with cols[3]:
                st.markdown(f"🕐 {f['caricato_il'][:10] if f['caricato_il'] else ''}")
            with cols[4]:
                if st.button("▶️", key=f"dash_open_{f['id']}", help="Apri"):
                    st.session_state.file_id = f["id"]
                    st.session_state.pagina = "vedi_file"
                    st.rerun()
    else:
        st.info("Nessun file caricato. Vai su **Carica Excel** per iniziare.")

# ── CARICA EXCEL ────────────────────────────────────────────────
elif st.session_state.pagina == "carica":
    st.markdown('<div class="page-title">📂 Carica file Excel</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Importa un file Excel nel database</div>', unsafe_allow_html=True)

    uploaded = st.file_uploader("Carica file Excel", type=["xlsx", "xls"], label_visibility="collapsed")
    if uploaded:
        try:
            contenuto = uploaded.read()
            df = leggi_excel_con_formule(uploaded)
            nome = uploaded.name
            colonne = list(df.columns) if not df.empty else []
            if not colonne:
                msg("Il file non ha colonne", "error")
            else:
                fid = db.salva_file_excel(st.session_state.utente["id"], nome,
                                          "Sheet1", colonne, len(df), contenuto)
                if not df.empty:
                    db.init_data_table(fid, df)
                else:
                    db.init_data_table_vuoto(fid, colonne)
                db.init_campi_config(fid, colonne)
                st.session_state.file_id = fid
                if len(df) > 0:
                    msg(f"✅ Importati {len(df)} record da '{nome}'")
                else:
                    msg(f"✅ Schema '{nome}' creato con {len(colonne)} colonne")
                st.session_state.pagina = "configura_campi"
                st.rerun()
        except Exception as e:
            msg(f"Errore: {e}", "error")

    st.divider()
    st.markdown("#### File caricati in precedenza")
    files = db.file_utente(st.session_state.utente["id"])
    if files:
        for f in files:
            cols = st.columns([3, 1, 1, 0.5])
            with cols[0]:
                st.markdown(f"**{f['nome_file']}**  \n{f['percorso']}")
            with cols[1]:
                st.markdown(f"📄 {f['foglio']}  ·  {f['righe']} righe")
            with cols[2]:
                if st.button("📌 Apri", key=f"carica_open_{f['id']}"):
                    st.session_state.file_id = f["id"]
                    st.session_state.pagina = "vedi_file"
                    st.rerun()
            with cols[3]:
                if st.button("🗑️", key=f"carica_del_{f['id']}", help="Elimina"):
                    db.elimina_file(f["id"])
                    msg(f"File '{f['nome_file']}' rimosso")
                    st.rerun()
    else:
        st.info("Nessun file caricato")

# ── VEDI FILE ───────────────────────────────────────────────────
elif st.session_state.pagina in ("vedi_file", "aggiungi_record", "modifica_record", "dettaglio_record"):
    fid = st.session_state.file_id
    if fid is None:
        msg("Nessun file selezionato", "warning")
        st.session_state.pagina = "dashboard"
        st.rerun()

    # Get file info
    files = db.file_utente(st.session_state.utente["id"])
    info = next((f for f in files if f["id"] == fid), None)
    if not info:
        msg("File non trovato", "error")
        st.session_state.pagina = "dashboard"
        st.rerun()

    col_tit, col_back = st.columns([3, 1])
    with col_tit:
        st.markdown(f'<div class="page-title">📊 {info["nome_file"]}</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="page-sub">{info["percorso"]}  ·  {info["foglio"]}  ·  {info["righe"]} righe</div>',
                     unsafe_allow_html=True)
    with col_back:
        if st.button("🏠  Dashboard", use_container_width=True):
            cambia_pagina("dashboard")

    db.migra_tabella(fid)
    df = db.carica_dati(fid)
    if df is None or df.empty:
        st.warning("Nessun dato disponibile")
        st.stop()

    colonne_dati = [c for c in df.columns if c not in ("id", "creato_da", "creato_il")]

    # Apply permissions
    u = st.session_state.utente
    puo_modificare = u["ruolo"] == "admin" or db.can_edit(u["id"], u["ruolo"], fid)
    if u["ruolo"] != "admin":
        perm = db.get_permesso_file(u["id"], fid)
        if perm:
            if perm.get("colonne_visibili"):
                colonne_dati = [c for c in colonne_dati if c in perm["colonne_visibili"]]
    # ── MODIFICA / AGGIUNGI ──
    if st.session_state.pagina == "modifica_record" or st.session_state.pagina == "aggiungi_record":
        if not puo_modificare:
            msg("Non hai il permesso di modificare i dati", "error")
            st.session_state.pagina = "vedi_file"
            st.rerun()
        is_new = st.session_state.pagina == "aggiungi_record"
        titolo = "➕ Nuovo record" if is_new else f"✏️ Modifica record #{st.session_state.modifica_id}"
        st.markdown(f"#### {titolo}")
        if st.session_state.formula_mode:
            st.caption("💡 Scrivi `=A1+B1`, `=SUM(A1:A10)`, `=IF(A1>5,\"Sì\",\"No\")` — le formule vengono valutate alla visualizzazione")

        campi_config = db.get_campi_config(fid)
        if not campi_config:
            campi_config = [{"id": 0, "nome_campo": c, "tipo_campo": "text", "obbligatorio": False, "opzioni": None, "mostra_modulo": True, "valore_predefinito": None} for c in colonne_dati]

        if is_new:
            campi_config = [cfg for cfg in campi_config if cfg.get("mostra_modulo", True)]

        rec = None
        if not is_new:
            rec_df = df[df["id"] == st.session_state.modifica_id]
            if not rec_df.empty:
                rec = rec_df.iloc[0]

        vals = {}
        errori = {}
        for cfg in campi_config:
            c = cfg["nome_campo"]
            if c not in colonne_dati:
                continue
            tipo = cfg["tipo_campo"]
            dv = rec[c] if rec is not None and pd.notna(rec[c]) else cfg.get("valore_predefinito")
            if dv is not None:
                if tipo == "number":
                    try:
                        dv = float(dv)
                    except:
                        dv = None
                elif tipo == "boolean":
                    dv = bool(dv)
            obbl = cfg["obbligatorio"]
            etichetta = f"{c} *" if obbl else c
            opts = cfg.get("opzioni") or []

            if tipo == "number":
                vals[c] = st.number_input(etichetta, value=float(dv) if dv is not None else None, step=0.01, format="%f", key=f"f_{c}")
            elif tipo == "date":
                vals[c] = st.date_input(etichetta, value=pd.to_datetime(dv) if dv is not None else None, key=f"f_{c}")
            elif tipo == "boolean":
                vals[c] = st.checkbox(etichetta, value=bool(dv) if dv is not None else False, key=f"f_{c}")
            elif tipo == "single_select":
                idx = opts.index(str(dv)) if dv is not None and str(dv) in opts else 0
                vals[c] = st.selectbox(etichetta, opts if opts else [""], index=idx, key=f"f_{c}")
            elif tipo == "multi_select":
                current = str(dv).split(",") if dv else []
                selected = st.multiselect(etichetta, opts if opts else [], default=[s for s in current if s in opts], key=f"f_{c}")
                vals[c] = ",".join(selected)
            elif tipo == "text_area":
                vals[c] = st.text_area(etichetta, value=str(dv) if dv is not None else "", key=f"f_{c}")
            else:
                vals[c] = st.text_input(etichetta, value=str(dv) if dv is not None else "", key=f"f_{c}")

            if obbl:
                v = vals[c]
                if v is None or (isinstance(v, str) and v.strip() == "") or (isinstance(v, float) and pd.isna(v)):
                    errori[c] = True

        if errori:
            msg("Compila tutti i campi obbligatori (*)", "warning")

        ca, cb = st.columns([1, 1])
        with ca:
            lbl = "💾 Aggiungi" if is_new else "💾 Salva modifiche"
            if st.button(lbl, use_container_width=True, type="primary"):
                if errori:
                    st.rerun()
                safe_vals = {}
                for c in colonne_dati:
                    v = vals.get(c)
                    if isinstance(v, str) and v.strip() == "":
                        safe_vals[c] = None
                    elif isinstance(v, bool):
                        safe_vals[c] = 1 if v else 0
                    else:
                        safe_vals[c] = v
                if is_new:
                    db.aggiungi_record(fid, safe_vals, st.session_state.utente["id"])
                    msg("✅ Record aggiunto")
                else:
                    db.aggiorna_record(fid, st.session_state.modifica_id, safe_vals)
                    msg(f"✅ Record #{st.session_state.modifica_id} aggiornato")
                st.session_state.pagina = "vedi_file"
                st.rerun()
        with cb:
            if st.button("Annulla", use_container_width=True):
                st.session_state.pagina = "vedi_file"
                st.rerun()
        st.stop()

    # ── VEDI DETTAGLIO ──
    if st.session_state.pagina == "dettaglio_record":
        rid = st.session_state.vedi_id
        rec_df = df[df["id"] == rid]
        if rec_df.empty:
            msg("Record non trovato", "error")
            st.session_state.pagina = "vedi_file"
            st.rerun()
        rec = rec_df.iloc[0]
        row_idx = df.index[df['id'] == rid][0]
        st.markdown(f"#### 👁️ Record #{rid}")
        for c in colonne_dati:
            v = rec[c]
            if st.session_state.formula_mode and isinstance(v, str) and v.startswith('='):
                v = evaluar_formula(v, df, row_idx, colonne_dati)
            v = "—" if pd.isna(v) else v
            st.markdown(f"**{c}:** {v}")
        st.divider()
        ce1, ce2, ce3 = st.columns([1, 1, 1])
        with ce1:
            if puo_modificare and st.button("✏️ Modifica", use_container_width=True):
                st.session_state.modifica_id = rid
                st.session_state.pagina = "modifica_record"
                st.rerun()
        with ce2:
            if puo_modificare and st.button("🗑️ Elimina", use_container_width=True):
                ok, err = db.elimina_record(fid, rid, st.session_state.utente["id"], st.session_state.utente["ruolo"])
                if ok:
                    msg(f"✅ Record #{rid} eliminato")
                    st.session_state.pagina = "vedi_file"
                else:
                    msg(err, "error")
                st.rerun()
        with ce3:
            if st.button("🔙 Indietro", use_container_width=True):
                st.session_state.pagina = "vedi_file"
                st.rerun()
        st.stop()

    # ── TOOLBAR ──
    tb = st.columns([1.5, 1, 1, 0.7, 0.7, 0.7, 0.7, 0.5, 0.5])
    with tb[0]:
        s = st.text_input("🔍 Cerca", value=st.session_state.ricerca,
                          placeholder="Cerca in tutto...", label_visibility="collapsed")
        if s != st.session_state.ricerca:
            st.session_state.ricerca = s
            st.rerun()
    with tb[1]:
        fc = st.selectbox("Colonna", [""] + colonne_dati, label_visibility="collapsed",
                          index=0 if not st.session_state.filtro_col
                          else (colonne_dati.index(st.session_state.filtro_col) + 1) if st.session_state.filtro_col in colonne_dati else 0)
        if fc != st.session_state.filtro_col:
            st.session_state.filtro_col = fc
            st.session_state.filtro_val = ""
            st.rerun()
    with tb[2]:
        if st.session_state.filtro_col:
            c = st.session_state.filtro_col
            unique_vals = sorted(df[c].dropna().unique())
            prev = st.session_state.get("filtro_val", "")
            fv = st.selectbox("Valore", ["(Tutti)"] + [str(v) for v in unique_vals],
                              index=0 if not prev or prev == "(Tutti)"
                              else ([str(v) for v in unique_vals].index(prev) + 1) if prev in [str(v) for v in unique_vals] else 0,
                              label_visibility="collapsed")
            if fv != st.session_state.get("filtro_val", ""):
                st.session_state.filtro_val = fv
                st.rerun()
    with tb[3]:
        if puo_modificare:
            if st.button("➕ Nuovo", use_container_width=True):
                st.session_state.pagina = "aggiungi_record"
                st.rerun()
    with tb[4]:
        if st.button("📥 Export", use_container_width=True, help="Esporta in Excel"):
            percorso = info["percorso"]
            if db.esporta_excel(fid, percorso):
                msg(f"✅ Dati esportati in '{percorso}'")
            else:
                msg("Errore durante l'esportazione", "error")
            st.rerun()
    with tb[5]:
        if st.button("⚙️ Campi", use_container_width=True, help="Configura campi"):
            st.session_state.pagina = "configura_campi"
            st.rerun()
    with tb[6]:
        v_next = {"tabella": "kanban", "kanban": "calendario", "calendario": "tabella"}
        v_lbl = {"tabella": "📋 Kanban", "kanban": "📅 Calendario", "calendario": "📊 Tabella"}
        cur = st.session_state.visuale
        if st.button(v_lbl.get(cur, "📊 Tabella"), use_container_width=True, help="Cambia visuale"):
            st.session_state.visuale = v_next.get(cur, "tabella")
            st.rerun()
    with tb[7]:
        f_label = "ƒx ON" if st.session_state.formula_mode else "ƒx OFF"
        f_help = "Disattiva formule" if st.session_state.formula_mode else "Attiva formule tipo Excel"
        if st.button(f_label, use_container_width=True, help=f_help, type="primary" if st.session_state.formula_mode else "secondary"):
            st.session_state.formula_mode = not st.session_state.formula_mode
            st.rerun()
    with tb[8]:
        if st.button("🔄", use_container_width=True, help="Ricarica dati"):
            try:
                buf, foglio_db = db.get_file_contenuto(fid)
                if buf:
                    import io
                    df_new = leggi_excel_con_formule(io.BytesIO(buf))
                else:
                    df_new = leggi_excel_con_formule(info["percorso"])
                db.init_data_table(fid, df_new)
                msg("✅ Dati ricaricati dal file Excel (formule preservate)")
            except Exception as e:
                msg(f"Errore: {e}", "error")
            st.rerun()

    # ── FILTER ──
    view_df = df
    if st.session_state.ricerca:
        q = st.session_state.ricerca.lower()
        mask = view_df[colonne_dati].astype(str).apply(lambda x: x.str.lower().str.contains(q, na=False)).any(axis=1)
        view_df = view_df[mask]

    if st.session_state.filtro_col and st.session_state.filtro_val and st.session_state.filtro_val != "(Tutti)":
        c = st.session_state.filtro_col
        v = st.session_state.filtro_val
        if c in view_df.columns:
            view_df = view_df[view_df[c].astype(str) == v]

    total = len(view_df)
    st.caption(f"{total} record{' (filtrati da ' + str(len(df)) + ' totali)' if total != len(df) else ''}")

    if view_df.empty:
        st.warning("Nessun record trovato")
        st.stop()

    # ── KANBAN / TABLE ──
    if st.session_state.visuale == "kanban":
        fcfg = db.get_file_config(fid)
        col_stato = fcfg.get("colonna_stato", "")
        col_titolo = fcfg.get("colonna_titolo", "carpeta")

        if not col_stato or col_stato not in colonne_dati:
            st.warning("⚠️ Colonna stato non configurata. Vai su **⚙️ Campi → 📋 Stati** per configurarla.")
            if st.button("⚙️ Configura stati"):
                st.session_state.pagina = "configura_stati"
                st.rerun()
            st.stop()

        stati = db.get_stati_config(fid, solo_kanban=True)
        if not stati:
            # fallback: all states
            stati = db.get_stati_config(fid)
        if not stati:
            vals = sorted(df[col_stato].dropna().unique())
            stati = [{"nome": str(v), "ordine": i, "colore": "#6B7280", "mostra_kanban": True} for i, v in enumerate(vals)]
            if not stati:
                stati = [{"nome": "Nessuno stato", "ordine": 0, "colore": "#6B7280", "mostra_kanban": True}]

        st.markdown(f'<div style="font-size:1.1rem;font-weight:600;margin-bottom:6px;">📋 Kanban — {col_stato}</div>', unsafe_allow_html=True)
        if st.button("⚙️ Configura stati"):
            st.session_state.pagina = "configura_stati"
            st.rerun()

        # Kanban columns — reorderable via ◀ ▶ buttons
        n_stati = len(stati)
        MAX_COLS = 6
        for batch_start in range(0, n_stati, MAX_COLS):
            batch = stati[batch_start:batch_start + MAX_COLS]
            cols = st.columns(len(batch))
            for i, stato in enumerate(batch):
                global_idx = batch_start + i
                with cols[i]:
                    nome = stato["nome"]
                    colore = stato.get("colore", "#6B7280")
                    records_in_col = view_df[view_df[col_stato].astype(str) == nome]

                    # Header with reorder buttons (nested columns)
                    hc = st.columns([0.1, 1, 0.1])
                    with hc[0]:
                        if global_idx > 0:
                            if st.button("◀", key=f"kl_{global_idx}", help=f"Sposta {nome} a sinistra"):
                                stati_list = [{"nome": s["nome"], "colore": s.get("colore", "#6B7280"), "mostra_kanban": s.get("mostra_kanban", True)} for s in stati]
                                prev = stati_list[global_idx - 1]
                                stati_list[global_idx - 1], stati_list[global_idx] = stati_list[global_idx], prev
                                db.save_stati_list(fid, stati_list)
                                st.rerun()
                    with hc[1]:
                        st.markdown(f'<div style="background:{colore};color:white;padding:6px;border-radius:6px;text-align:center;font-weight:600;font-size:0.8rem;">{nome} <span style="font-weight:400;">({len(records_in_col)})</span></div>', unsafe_allow_html=True)
                    with hc[2]:
                        if global_idx < n_stati - 1:
                            if st.button("▶", key=f"kr_{global_idx}", help=f"Sposta {nome} a destra"):
                                stati_list = [{"nome": s["nome"], "colore": s.get("colore", "#6B7280"), "mostra_kanban": s.get("mostra_kanban", True)} for s in stati]
                                nxt = stati_list[global_idx + 1]
                                stati_list[global_idx], stati_list[global_idx + 1] = nxt, stati_list[global_idx]
                                db.save_stati_list(fid, stati_list)
                                st.rerun()

                    # Cards
                    for _, rec in records_in_col.iterrows():
                        rid = int(rec["id"])
                        titolo = str(rec.get(col_titolo, "")) if col_titolo and col_titolo in rec and pd.notna(rec.get(col_titolo)) else f"#{rid}"
                        with st.container():
                            st.markdown(f'<div class="kanban-card"><span style="font-weight:600;font-size:0.85rem;color:#1E3A5F;">{titolo}</span><br><span style="color:#9CA3AF;font-size:0.7rem;">#{rid}</span></div>', unsafe_allow_html=True)
                            other_stati = [s["nome"] for s in stati if s["nome"] != nome]
                            if other_stati:
                                key = f"mv_{rid}_{nome.replace(' ','_')}"
                                new_stato = st.selectbox("", ["—"] + other_stati, key=key, label_visibility="collapsed")
                                if new_stato and new_stato != "—":
                                    trans = db.get_transizioni(fid)
                                    for t in trans:
                                        if t["stato_da"] == nome and t["stato_a"] == new_stato:
                                            action = t["azione_tipo"]
                                            col = t["colonna_destinazione"]
                                            val = t["valore"]
                                            if action == "set_data" and col:
                                                from datetime import date
                                                db.aggiorna_record(fid, rid, {col: str(date.today())})
                                            elif action == "set_timestamp" and col:
                                                from datetime import datetime
                                                db.aggiorna_record(fid, rid, {col: str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))})
                                            elif action == "set_valore" and col and val:
                                                db.aggiorna_record(fid, rid, {col: val})
                                    db.aggiorna_record(fid, rid, {col_stato: new_stato})
                                    st.toast(f"📦 #{rid} → {new_stato}")
                                    st.rerun()
        st.stop()

    # ── CALENDARIO ──
    if st.session_state.visuale == "calendario":
        fcfg = db.get_file_config(fid)
        col_data = fcfg.get("colonna_data_calendario", "")
        col_titolo = fcfg.get("colonna_titolo", "")

        if not col_data or col_data not in colonne_dati:
            st.warning("⚠️ Colonna data non configurata. Vai su **⚙️ Campi → 📋 Stati** per selezionare la colonna data per il calendario.")
            st.stop()

        oggi = date.today()
        events = []
        for _, rec in view_df.iterrows():
            val = rec.get(col_data)
            if pd.notna(val):
                try:
                    start = str(val)[:10]
                    titolo = str(rec.get(col_titolo, "")) if col_titolo and col_titolo in rec and pd.notna(rec.get(col_titolo)) else f"#{int(rec['id'])}"
                    ev = {"id": str(int(rec["id"])), "title": titolo, "start": start}
                    try:
                        if datetime.strptime(start, "%Y-%m-%d").date() < oggi:
                            ev["backgroundColor"] = "#EF4444"
                            ev["borderColor"] = "#DC2626"
                            ev["textColor"] = "#FFFFFF"
                            ev["title"] = "⚠ " + titolo
                    except Exception:
                        pass
                    events.append(ev)
                except Exception:
                    pass

        if not events:
            st.info("Nessun evento con data da mostrare nel calendario.")
            st.stop()

        from streamlit_calendar import calendar

        cal_options = {
            "editable": True,
            "selectable": True,
            "headerToolbar": {
                "left": "today prev,next",
                "center": "title",
                "right": "dayGridMonth,dayGridWeek,dayGridDay",
            },
            "initialView": "dayGridMonth",
            "height": 650,
        }

        cal_result = calendar(events=events, options=cal_options, key=f"cal_{fid}")

        if cal_result and "eventDrop" in cal_result:
            drop = cal_result["eventDrop"]["event"]
            drop_tag = f"{drop['id']}_{drop['start']}"
            if st.session_state.get("_cal_drop") != drop_tag:
                st.session_state._cal_drop = drop_tag
                db.aggiorna_record(fid, int(drop["id"]), {col_data: drop["start"][:10]})
                st.toast(f"📅 #{drop['id']} spostato al {drop['start'][:10]}")
                st.rerun()

        st.divider()
        st.caption("Trascina un evento per cambiarne la data.")

        st.stop()

    # ── TABLE ──
    display_cols = colonne_dati
    if st.session_state.formula_mode:
        view_df_display = aplicar_formulas(view_df, display_cols)
    else:
        view_df_display = view_df
    ev = st.dataframe(
        view_df_display[["id"] + display_cols],
        use_container_width=True,
        hide_index=True,
        column_config={"id": "ID"},
        on_select="rerun",
        selection_mode="single-row",
    )

    # ── ACTIONS ──
    if ev.selection and ev.selection.rows:
        sel_idx = view_df.index[ev.selection.rows[0]]
        sel = view_df.loc[sel_idx]
        rid = int(sel["id"])

        st.divider()
        st.markdown(f"**Record #{rid}** selezionato")
        # preview (show computed values if formula mode is on)
        if st.session_state.formula_mode:
            sel_display = aplicar_formulas(pd.DataFrame([sel]), display_cols).iloc[0]
        else:
            sel_display = sel
        preview = sel_display[display_cols[:min(4, len(display_cols))]]
        for c, v in preview.items():
            v = "—" if pd.isna(v) else v
            st.markdown(f"**{c}:** {v}")
        if len(display_cols) > 4:
            st.caption(f"+ {len(display_cols) - 4} campi")

        ac1, ac2, ac3 = st.columns([1, 1, 8])
        with ac1:
            if st.button("👁️ Vedi", use_container_width=True):
                st.session_state.vedi_id = rid
                st.session_state.pagina = "dettaglio_record"
                st.rerun()
        with ac2:
            if puo_modificare and st.button("✏️ Modifica", use_container_width=True):
                st.session_state.modifica_id = rid
                st.session_state.pagina = "modifica_record"
                st.rerun()
        with ac3:
            if puo_modificare and st.button("🗑️ Elimina", use_container_width=True):
                ok, err = db.elimina_record(fid, rid, st.session_state.utente["id"], st.session_state.utente["ruolo"])
                if ok:
                    msg(f"✅ Record #{rid} eliminato")
                else:
                    msg(err, "error")
                st.rerun()
    else:
        st.info("👆 Seleziona una riga per vedere le azioni")

# ── CONFIGURA CAMPI ─────────────────────────────────────────────
elif st.session_state.pagina == "configura_campi":
    fid = st.session_state.file_id
    if fid is None:
        msg("Nessun file selezionato", "warning")
        st.session_state.pagina = "dashboard"
        st.rerun()

    if "modifiche_pendenti" not in st.session_state:
        st.session_state.modifiche_pendenti = []

    st.markdown(f'<div class="page-title">⚙️ Configura campi</div>', unsafe_allow_html=True)
    col_back, col_stati, col_add = st.columns([1, 1, 2])
    with col_back:
        if st.button("🔙 Torna al file", use_container_width=False):
            st.session_state.modifiche_pendenti = []
            st.session_state.pagina = "vedi_file"
            st.rerun()
    with col_stati:
        if st.button("📋 Stati", use_container_width=False, help="Configura stati Kanban"):
            st.session_state.pagina = "configura_stati"
            st.rerun()
    with col_add:
        with st.expander("➕ Aggiungi colonna"):
            with st.form("nuova_colonna", border=False):
                c_nome = st.text_input("Nome colonna", placeholder="Nuovo campo")
                c_tipo = st.selectbox("Tipo", ["text", "text_area", "number", "date", "single_select", "multi_select", "boolean"])
                if st.form_submit_button("➕ Accoda", use_container_width=True):
                    if c_nome.strip():
                        st.session_state.modifiche_pendenti.append({"tipo": "aggiungi", "nome": c_nome.strip(), "tipo_campo": c_tipo})
                        st.markdown(f'<span style="color:#10B981;">✅ "{c_nome.strip()}" in coda</span>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<span style="color:#EF4444;">Inserisci un nome</span>', unsafe_allow_html=True)
    st.divider()

    campi = db.get_campi_config(fid)
    if not campi:
        df = db.carica_dati(fid)
        if df is not None:
            db.init_campi_config(fid, [c for c in df.columns if c != "id"])
            campi = db.get_campi_config(fid)
    if not campi:
        msg("Nessuna colonna trovata", "error")
        st.session_state.pagina = "vedi_file"
        st.rerun()

    # ── PENDING CHANGES PANEL ──
    if st.session_state.modifiche_pendenti:
        pendenti = st.session_state.modifiche_pendenti
        n_queue = sum(1 for m in pendenti if m["tipo"] != "elimina")
        n_del = sum(1 for m in pendenti if m["tipo"] == "elimina")
        with st.container():
            st.markdown(f"#### 📋 Modifiche in coda ({len(pendenti)})")
            for i, m in enumerate(pendenti):
                if m["tipo"] == "aggiorna":
                    label = f"✏️ **{m['campo_old']}** — tipo: `{m['tipo_campo']}`, obbl: {'✓' if m['obbl'] else '✗'}, form: {'✓' if m['mostra'] else '✗'}"
                elif m["tipo"] == "rinomina":
                    label = f"🔤 **{m['vecchio']}** → **{m['nuovo']}**"
                elif m["tipo"] == "elimina":
                    label = f"🗑️ **{m['nome']}**"
                elif m["tipo"] == "aggiungi":
                    label = f"➕ **{m['nome']}** (`{m['tipo_campo']}`)"
                else:
                    label = str(m)
                st.markdown(f"- {label}")
            ca1, ca2 = st.columns([1, 1])
            with ca1:
                if st.button("✅ Applica tutte", use_container_width=True, type="primary"):
                    for m in pendenti:
                        try:
                            if m["tipo"] == "aggiorna":
                                db.aggiorna_campo_config(m["config_id"], tipo_campo=m["tipo_campo"], obbligatorio=m["obbl"], opzioni=m.get("opzioni"), mostra_modulo=m["mostra"], valore_predefinito=m.get("default"))
                            elif m["tipo"] == "rinomina":
                                db.rinomina_colonna(fid, m["vecchio"], m["nuovo"])
                            elif m["tipo"] == "elimina":
                                db.elimina_colonna(fid, m["config_id"], m["nome"])
                            elif m["tipo"] == "aggiungi":
                                db.aggiungi_colonna(fid, m["nome"], m["tipo_campo"])
                        except Exception:
                            pass
                    st.session_state.modifiche_pendenti = []
                    msg("✅ Modifiche applicate")
                    st.rerun()
            with ca2:
                if st.button("🗑️ Annulla tutto", use_container_width=True):
                    st.session_state.modifiche_pendenti = []
                    st.rerun()
        st.divider()

    # Header
    hdr = st.columns([1.5, 1.3, 0.6, 0.6, 1.5, 0.5, 0.4])
    for h, label in zip(hdr, ["Campo", "Tipo", "Obbl.", "Form", "Opzioni/Default", "", ""]):
        with h:
            st.markdown(f"**{label}**" if label else "")

    for campo in campi:
        with st.container():
            cols = st.columns([1.5, 1.3, 0.6, 0.6, 1.5, 0.5, 0.4])
            with cols[0]:
                nuovo_nome = st.text_input("Nome", value=campo["nome_campo"], key=f"nome_{campo['id']}", label_visibility="collapsed")
            with cols[1]:
                tipo = st.selectbox(
                    "Tipo",
                    ["text", "text_area", "number", "date", "single_select", "multi_select", "boolean"],
                    index=["text", "text_area", "number", "date", "single_select", "multi_select", "boolean"].index(campo["tipo_campo"]) if campo["tipo_campo"] in ["text", "text_area", "number", "date", "single_select", "multi_select", "boolean"] else 0,
                    key=f"tipo_{campo['id']}",
                    label_visibility="collapsed",
                )
            with cols[2]:
                obbl = st.checkbox("Obbligatorio", value=campo["obbligatorio"], key=f"obbl_{campo['id']}", label_visibility="collapsed")
            with cols[3]:
                mostra = st.checkbox("Mostra nel form", value=campo["mostra_modulo"], key=f"mostra_{campo['id']}", label_visibility="collapsed")
            with cols[4]:
                is_select = tipo in ("single_select", "multi_select")
                if is_select:
                    opts_key = f"opts_{campo['id']}"
                    csv_key = f"csv_{campo['id']}"
                    current_opts = ", ".join(campo.get("opzioni", [])) if campo.get("opzioni") else ""
                    csv_file = st.file_uploader("CSV", type=["csv"], key=csv_key, label_visibility="collapsed")
                    if csv_file:
                        try:
                            df_csv = pd.read_csv(csv_file)
                            if not df_csv.empty:
                                vals = df_csv.iloc[:, 0].dropna().astype(str).tolist()
                                st.session_state[opts_key] = ", ".join(vals)
                        except Exception:
                            pass
                    opts_str = st.text_input("Opzioni", value=st.session_state.get(opts_key, current_opts),
                                             key=opts_key, label_visibility="collapsed",
                                             placeholder="Opz1, Opz2, ...")
                else:
                    dv = campo.get("valore_predefinito") or ""
                    opts_str = st.text_input("Default", value=dv,
                                             key=f"def_{campo['id']}", label_visibility="collapsed",
                                             placeholder="Valore default")
            with cols[5]:
                if st.button("📋", key=f"queue_{campo['id']}", help="Accoda modifica"):
                    new_opzioni = None
                    new_default = None
                    if is_select:
                        if opts_str and opts_str.strip():
                            new_opzioni = [o.strip() for o in opts_str.split(",")]
                        else:
                            df_cfg = db.carica_dati(fid)
                            if df_cfg is not None and campo['nome_campo'] in df_cfg.columns:
                                vals = df_cfg[campo['nome_campo']].dropna().unique()
                                new_opzioni = sorted([str(v) for v in vals if v != "" and str(v).strip() != ""])
                                if not new_opzioni:
                                    new_opzioni = [" "]
                    else:
                        new_default = opts_str.strip() if opts_str and opts_str.strip() else None
                    st.session_state.modifiche_pendenti.append({
                        "tipo": "aggiorna", "config_id": campo["id"],
                        "tipo_campo": tipo, "obbl": obbl, "mostra": mostra,
                        "opzioni": new_opzioni, "default": new_default,
                        "campo_old": campo["nome_campo"],
                    })
                    if nuovo_nome.strip() and nuovo_nome.strip() != campo["nome_campo"]:
                        st.session_state.modifiche_pendenti.append({
                            "tipo": "rinomina", "config_id": campo["id"],
                            "vecchio": campo["nome_campo"], "nuovo": nuovo_nome.strip(),
                        })
                    st.toast(f"📋 '{campo['nome_campo']}' accodato ({len(st.session_state.modifiche_pendenti)})")
            with cols[6]:
                if st.button("🗑️", key=f"del_cfg_{campo['id']}", help="Elimina"):
                    st.session_state.modifiche_pendenti.append({
                        "tipo": "elimina", "config_id": campo["id"], "nome": campo["nome_campo"],
                    })
                    st.toast(f"🗑️ '{campo['nome_campo']}' in coda per eliminazione")
            st.divider()

# ── CONFIGURA STATI (Kanban) ────────────────────────────────────
elif st.session_state.pagina == "configura_stati":
    fid = st.session_state.file_id
    if fid is None:
        msg("Nessun file selezionato", "warning")
        st.session_state.pagina = "dashboard"
        st.rerun()

    st.markdown(f'<div class="page-title">📋 Configura stati</div>', unsafe_allow_html=True)

    col_back, _ = st.columns([1, 3])
    with col_back:
        if st.button("🔙 Torna al file", use_container_width=False):
            st.session_state.pagina = "vedi_file"
            st.rerun()
    st.divider()

    fcfg = db.get_file_config(fid)
    campi_disponibili = []
    df_cfg = db.carica_dati(fid)
    if df_cfg is not None:
        campi_disponibili = [c for c in df_cfg.columns if c not in ("id", "creato_da", "creato_il")]

    col_stato = st.selectbox("Colonna stato", [""] + campi_disponibili,
                             index=0 if not fcfg.get("colonna_stato") else (campi_disponibili.index(fcfg["colonna_stato"]) + 1) if fcfg["colonna_stato"] in campi_disponibili else 0,
                             help="Colonna che contiene lo stato di ogni record")
    col_titolo = st.selectbox("Colonna titolo (card)", [""] + campi_disponibili,
                              index=0 if not fcfg.get("colonna_titolo") else (campi_disponibili.index(fcfg["colonna_titolo"]) + 1) if fcfg["colonna_titolo"] in campi_disponibili else 0,
                              help="Colonna usata come titolo della card nel Kanban e calendario")
    col_data_cal = st.selectbox("Colonna data (calendario)", [""] + campi_disponibili,
                                index=0 if not fcfg.get("colonna_data_calendario") else (campi_disponibili.index(fcfg["colonna_data_calendario"]) + 1) if fcfg["colonna_data_calendario"] in campi_disponibili else 0,
                                help="Colonna data usata per il calendario con trascinamento")
    if st.button("💾 Salva colonne", use_container_width=False):
        db.save_file_config(fid, col_stato, col_titolo, col_data_cal)
        msg("✅ Colonne salvate")
        st.rerun()

    st.divider()

    if not col_stato:
        st.info("Seleziona prima la colonna 'Stato'")
        st.stop()

    # Stati
    st.markdown("#### Ordine e colori degli stati")
    stati = db.get_stati_config(fid)
    if not stati:
        # auto-populate from data
        if df_cfg is not None and col_stato in df_cfg.columns:
            vals = sorted(df_cfg[col_stato].dropna().unique())
            stati = [{"id": 0, "nome": str(v), "ordine": i, "colore": "#6B7280", "mostra_kanban": True} for i, v in enumerate(vals)]

    stati_nomi = [s["nome"] for s in stati]
    colori_default = ["#1E3A5F", "#059669", "#D97706", "#DC2626", "#7C3AED", "#0891B2", "#BE185D", "#4B5563"]

    with st.form("stati_form", border=False):
        nuovi_stati = []
        for i in range(max(len(stati), 2)):
            default_nome = stati[i]["nome"] if i < len(stati) else ""
            default_colore = stati[i].get("colore", colori_default[i % len(colori_default)]) if i < len(stati) else colori_default[i % len(colori_default)]
            default_mostra = stati[i].get("mostra_kanban", True) if i < len(stati) else True
            sc1, sc2, sc3, sc4 = st.columns([2, 1, 0.5, 0.5])
            with sc1:
                nome_s = st.text_input(f"Stato {i+1}", value=default_nome, key=f"st_nome_{i}", label_visibility="collapsed", placeholder="Nome stato")
            with sc2:
                colore_s = st.color_picker("Colore", value=default_colore, key=f"st_col_{i}", label_visibility="collapsed")
            with sc3:
                st.markdown(f"**#{i+1}**" if default_nome or i == 0 else "")
            with sc4:
                mostra_s = st.checkbox("Kanban", value=default_mostra, key=f"st_mostra_{i}", label_visibility="collapsed", help="Mostra in Kanban")
            if nome_s and nome_s.strip():
                nuovi_stati.append({"nome": nome_s.strip(), "colore": colore_s, "mostra_kanban": mostra_s})
        if st.form_submit_button("💾 Salva stati", use_container_width=True, type="primary"):
            if nuovi_stati:
                db.save_stati_list(fid, nuovi_stati)
                msg(f"✅ {len(nuovi_stati)} stati salvati")
                st.rerun()
            else:
                msg("Inserisci almeno uno stato", "warning")

    st.divider()
    st.markdown("#### ⚡ Transizioni (azioni al cambio stato)")
    st.caption("Esempio: quando sposti da 'Da iniziare' a 'Pasc evaso', imposta la data odierna nella colonna 'Data inizio'")

    trans = db.get_transizioni(fid)
    with st.form("nuova_transizione", border=False):
        tutti_stati = list(set([s["nome"] for s in stati] + [t["stato_da"] for t in trans] + [t["stato_a"] for t in trans]))
        if not tutti_stati:
            tutti_stati = stati_nomi
        td = st.selectbox("Da stato", [""] + tutti_stati, key="trans_da")
        ta = st.selectbox("A stato", [""] + tutti_stati, key="trans_a")
        azione = st.selectbox("Azione", ["", "set_data", "set_timestamp", "set_valore"], key="trans_az",
                              format_func=lambda x: {"": "Seleziona azione", "set_data": "Imposta data odierna", "set_timestamp": "Imposta timestamp", "set_valore": "Imposta valore"}.get(x, x))
        if azione == "set_valore":
            ac = st.selectbox("Colonna destinazione", [""] + campi_disponibili, key="trans_col_val")
            av = st.text_input("Valore", key="trans_val")
        elif azione:
            ac = st.selectbox("Colonna destinazione", [""] + campi_disponibili, key="trans_col")
            av = ""
        else:
            ac = ""
            av = ""
        if st.form_submit_button("➕ Aggiungi transizione", use_container_width=True, type="primary"):
            if td and ta and azione and ac:
                db.save_transizione(fid, td, ta, azione, ac, av)
                msg(f"✅ Transizione: {td} → {ta}")
                st.rerun()
            else:
                msg("Compila Da, A, Azione e Colonna", "warning")

    if trans:
        st.markdown("**Transizioni esistenti:**")
        for t in trans:
            ct1, ct2, ct3 = st.columns([2, 3, 0.5])
            with ct1:
                az_label = {"set_data": "📅 Data", "set_timestamp": "⏰ Timestamp", "set_valore": "📝 Valore"}.get(t["azione_tipo"], t["azione_tipo"])
                st.markdown(f"**{t['stato_da']}** → **{t['stato_a']}**: {az_label} in `{t['colonna_destinazione']}`" + (f" = '{t['valore']}'" if t['valore'] else ""))
            with ct2:
                st.caption(f"id: {t['id']}")
            with ct3:
                if st.button("🗑️", key=f"del_trans_{t['id']}", help="Elimina transizione"):
                    db.elimina_transizione(t["id"])
                    msg("Transizione eliminata")
                    st.rerun()
    else:
        st.caption("Nessuna transizione configurata")

# ── GESTIONE UTENTI (admin only) ────────────────────────────────
elif st.session_state.pagina == "utenti":
    if st.session_state.utente["ruolo"] != "admin":
        msg("Accesso non autorizzato", "error")
        cambia_pagina("dashboard")

    st.markdown('<div class="page-title">👥 Gestione Utenti</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Crea, modifica password e assegna permessi file</div>', unsafe_allow_html=True)

    with st.expander("➕ Crea nuovo utente", expanded=False):
        with st.form("nuovo_utente"):
            col1, col2 = st.columns(2)
            with col1:
                nome = st.text_input("Nome completo")
                email_nuovo = st.text_input("Email")
            with col2:
                pw = st.text_input("Password", type="password")
                ruolo = st.selectbox("Ruolo", ["utente", "admin"])
            if st.form_submit_button("Crea utente", use_container_width=True, type="primary"):
                if nome and email_nuovo and pw:
                    ok, msg_text = db.crea_utente(email_nuovo, nome, pw, ruolo)
                    if ok:
                        msg(msg_text)
                        st.rerun()
                    else:
                        msg(msg_text, "error")
                else:
                    msg("Compila tutti i campi", "warning")

    st.markdown("#### Cambia password")
    with st.form("cambia_pw"):
        sel_user = st.selectbox("Utente", [f"{u['nome']} ({u['email']})" for u in db.lista_utenti()], key="pw_user")
        nuova_pw = st.text_input("Nuova password", type="password")
        if st.form_submit_button("Aggiorna password", use_container_width=True, type="primary"):
            if nuova_pw:
                uid = [u for u in db.lista_utenti() if f"{u['nome']} ({u['email']})" == sel_user][0]["id"]
                db.cambia_password(uid, nuova_pw)
                msg("Password aggiornata")
                st.rerun()
            else:
                msg("Inserisci la nuova password", "warning")

    st.divider()
    st.markdown("#### Elenco utenti")
    utenti = db.lista_utenti()
    for ut in utenti:
        with st.expander(f"**{ut['nome']}** — {ut['email']} (`{ut['ruolo']}`)", expanded=False):
            if ADMIN_EMAIL and ut["email"] == ADMIN_EMAIL:
                st.caption("Admin principale — non modificabile")
                continue
            col_pw, col_file = st.columns([1, 2])
            with col_pw:
                st.markdown("**Password**")
                with st.form(key=f"pw_{ut['id']}", border=False):
                    npw = st.text_input("Nuova password", type="password", key=f"npw_{ut['id']}", label_visibility="collapsed")
                    if st.form_submit_button("Salva password", use_container_width=True):
                        if npw:
                            db.cambia_password(ut["id"], npw)
                            msg("Password cambiata")
                            st.rerun()
            with col_file:
                st.markdown("**Permessi file**")
                all_files = db.get_all_files()
                for f in all_files:
                    perm = db.get_permesso_file(ut["id"], f["id"])
                    has_perm = perm is not None
                    kb = st.checkbox(f["nome_file"], value=has_perm, key=f"fp_{ut['id']}_{f['id']}")
                    if kb:
                        campi = [c["nome_campo"] for c in db.get_campi_config(f["id"])] or []
                        col_vis = st.multiselect("Colonne visibili", campi,
                                                  default=perm["colonne_visibili"] if perm and perm.get("colonne_visibili") else campi,
                                                  key=f"fp_col_{ut['id']}_{f['id']}", label_visibility="collapsed")
                        if col_vis:
                            non_vis = [c for c in campi if c not in col_vis]
                            if non_vis:
                                st.caption(f"Nascoste: {', '.join(non_vis)}")
                        mod_perm = st.checkbox("Permetti modifica", value=perm["modifica"] if perm and perm.get("modifica") else False,
                                               key=f"fp_mod_{ut['id']}_{f['id']}", help="Consenti all'utente di aggiungere/modificare/eliminare record")
                        with st.form(key=f"fp_save_{ut['id']}_{f['id']}", border=False):
                            if st.form_submit_button("Salva permessi", use_container_width=True):
                                db.set_permesso_file(ut["id"], f["id"], col_vis if col_vis else None, modifica=mod_perm)
                                msg(f"Permessi aggiornati per {ut['nome']} su {f['nome_file']}")
                                st.rerun()
                    elif has_perm:
                        db.elimina_permesso_file(ut["id"], f["id"])
                        msg(f"Permesso rimosso per {ut['nome']} su {f['nome_file']}")
                        st.rerun()
            st.divider()
            if st.button(f"🗑️ Elimina utente {ut['nome']}", key=f"del_user_{ut['id']}", use_container_width=True):
                db.elimina_utente(ut["id"])
                msg(f"Utente {ut['nome']} eliminato")
                st.rerun()
